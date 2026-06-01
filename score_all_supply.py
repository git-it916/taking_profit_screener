#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
전종목_수급.xlsx 기반 스코어링

입력:
    C:/Users/10845/Documents/quant_project/전종목_수급.xlsx

출력:
    C:/Users/10845/Documents/quant_project/[오후] whole-stock-2/전종목_수급_스코어링_YYYYMMDD.xlsx
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd


BASE_DIR = Path(r"C:\Users\10845\Documents\quant_project")
INPUT_FILE = BASE_DIR / "전종목_수급.xlsx"
OUTPUT_DIR = BASE_DIR / "[오후] whole-stock-2"
SOURCE_SHEET = "Sheet1"

PRICE_WEIGHT = 0.35
SUPPLY_WEIGHT = 0.45
VALUE_WEIGHT = 0.20


ITEM_MAP = {
    "종가(원)": "close",
    "거래대금 (5일 평균)(원)": "turnover_avg_5d",
    "거래대금 (20일 평균)(원)": "turnover_avg_20d",
    "순매수대금(기관계)(5일합산)(만원)": "inst_net_5d",
    "순매수대금(기관계)(20일합산)(만원)": "inst_net_20d",
    "순매수대금(외국인계)(5일합산)(만원)": "foreign_net_5d",
    "순매수대금(외국인계)(20일합산)(만원)": "foreign_net_20d",
}


def normalize_score(series: pd.Series, higher_is_better: bool = True) -> pd.Series:
    series = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
    if series.notna().sum() == 0:
        return pd.Series(50.0, index=series.index)

    pct = series.rank(pct=True, na_option="keep")
    score = pct * 100 if higher_is_better else (1 - pct) * 100
    return score.fillna(50).round(1)


def choose_data_sheet(path: Path) -> str:
    xl = pd.ExcelFile(path)
    if SOURCE_SHEET in xl.sheet_names:
        return SOURCE_SHEET

    best_sheet = xl.sheet_names[0]
    best_count = -1

    for sheet in xl.sheet_names:
        preview = pd.read_excel(path, sheet_name=sheet, header=None, nrows=13)
        if len(preview) <= 8:
            continue
        symbols = preview.iloc[8].astype(str)
        count = symbols.str.match(r"^A\d{6}$").sum()
        if count > best_count:
            best_count = int(count)
            best_sheet = sheet

    return best_sheet


def load_workbook_data(path: Path) -> tuple[pd.DataFrame, str]:
    sheet = choose_data_sheet(path)
    raw = pd.read_excel(path, sheet_name=sheet, header=None)

    symbols = raw.iloc[8]
    names = raw.iloc[9]
    item_names = raw.iloc[12]

    data = raw.iloc[14:].copy()
    data = data.rename(columns={0: "date"})
    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    data = data.dropna(subset=["date"]).reset_index(drop=True)

    frames = []
    unique_symbols = []
    for col, symbol in symbols.items():
        symbol = str(symbol).strip()
        if not pd.Series([symbol]).str.match(r"^A\d{6}$").iloc[0]:
            continue

        item = str(item_names.get(col, "")).strip()
        field = ITEM_MAP.get(item)
        if field is None:
            continue

        name = names.get(col)
        unique_symbols.append(symbol)
        frame = pd.DataFrame(
            {
                "date": data["date"],
                "code": symbol,
                "name": "" if pd.isna(name) else str(name),
                "field": field,
                "value": pd.to_numeric(data[col], errors="coerce"),
            }
        )
        frames.append(frame)

    if not frames:
        raise RuntimeError("전종목_수급.xlsx에서 스코어링 가능한 종목 데이터를 찾지 못했습니다.")

    long_df = pd.concat(frames, ignore_index=True)
    wide = (
        long_df.pivot_table(
            index=["date", "code", "name"],
            columns="field",
            values="value",
            aggfunc="first",
        )
        .reset_index()
        .sort_values(["code", "date"])
    )
    wide.columns.name = None

    print(f"사용 시트: {sheet}")
    print(f"파싱 종목 수: {pd.Series(unique_symbols).nunique():,}개")
    print(f"데이터 기간: {wide['date'].min().strftime('%Y-%m-%d')} ~ {wide['date'].max().strftime('%Y-%m-%d')}")
    return wide, sheet


def add_time_series_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.sort_values(["code", "date"]).copy()

    grouped = df.groupby("code", group_keys=False)
    df["ret_5d"] = grouped["close"].pct_change(5) * 100
    df["ret_20d"] = grouped["close"].pct_change(20) * 100
    df["ma5"] = grouped["close"].transform(lambda s: s.rolling(5, min_periods=3).mean())
    df["ma10"] = grouped["close"].transform(lambda s: s.rolling(10, min_periods=5).mean())
    df["ma20"] = grouped["close"].transform(lambda s: s.rolling(20, min_periods=10).mean())

    # 수급은 만원 단위, 거래대금은 원 단위입니다.
    df["inst_ratio_5d"] = (df["inst_net_5d"] * 10_000) / (df["turnover_avg_5d"] * 5) * 100
    df["foreign_ratio_5d"] = (df["foreign_net_5d"] * 10_000) / (df["turnover_avg_5d"] * 5) * 100
    df["inst_ratio_20d"] = (df["inst_net_20d"] * 10_000) / (df["turnover_avg_20d"] * 20) * 100
    df["foreign_ratio_20d"] = (df["foreign_net_20d"] * 10_000) / (df["turnover_avg_20d"] * 20) * 100

    df["inst_positive_rate_20d"] = grouped["inst_ratio_5d"].transform(
        lambda s: (s > 0).rolling(20, min_periods=5).mean() * 100
    )
    df["foreign_positive_rate_20d"] = grouped["foreign_ratio_5d"].transform(
        lambda s: (s > 0).rolling(20, min_periods=5).mean() * 100
    )
    df["turnover_rvol"] = df["turnover_avg_5d"] / df["turnover_avg_20d"]
    return df.replace([np.inf, -np.inf], np.nan)


def latest_scoring_frame(df: pd.DataFrame) -> pd.DataFrame:
    latest_date = df["date"].max()
    latest = df[df["date"] == latest_date].copy()

    latest["above_ma10"] = latest["close"] > latest["ma10"]
    latest["ma5_above_ma10"] = latest["ma5"] > latest["ma10"]
    latest["both_supply_positive"] = (latest["inst_ratio_5d"] > 0) & (latest["foreign_ratio_5d"] > 0)

    latest["price_score"] = (
        normalize_score(latest["ret_20d"]).clip(2, 98) * 0.40
        + normalize_score(latest["ret_5d"]).clip(2, 98) * 0.30
        + latest["above_ma10"].astype(float) * 100 * 0.20
        + latest["ma5_above_ma10"].astype(float) * 100 * 0.10
    ).round(1)

    latest["supply_score"] = (
        normalize_score(latest["foreign_ratio_5d"]) * 0.25
        + normalize_score(latest["foreign_ratio_20d"]) * 0.15
        + normalize_score(latest["foreign_positive_rate_20d"]) * 0.10
        + normalize_score(latest["inst_ratio_5d"]) * 0.25
        + normalize_score(latest["inst_ratio_20d"]) * 0.15
        + normalize_score(latest["inst_positive_rate_20d"]) * 0.05
        + latest["both_supply_positive"].astype(float) * 100 * 0.05
    ).round(1)

    latest["value_score"] = (
        normalize_score(latest["turnover_rvol"]) * 0.70
        + normalize_score(latest["turnover_avg_5d"]) * 0.30
    ).round(1)

    latest["total_score"] = (
        latest["price_score"] * PRICE_WEIGHT
        + latest["supply_score"] * SUPPLY_WEIGHT
        + latest["value_score"] * VALUE_WEIGHT
    ).round(1)

    latest["grade"] = pd.cut(
        latest["total_score"],
        bins=[float("-inf"), 30, 50, 65, 80, float("inf")],
        labels=["F", "D", "C", "B", "A"],
        right=False,
    ).astype(str)

    return latest.sort_values("total_score", ascending=False, na_position="last")


def _format_table_sheet(ws, freeze_row: int = 2) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    navy = "1F3864"
    white = "FFFFFF"
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    ws.freeze_panes = f"A{freeze_row}"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True, color=white, name="맑은 고딕", size=9)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter][: min(ws.max_row, 80)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 24)


def _create_report_sheet(writer, final: pd.DataFrame, scored: pd.DataFrame, ref_date: str) -> None:
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = writer.book
    ws = wb.create_sheet("전종목 보고서", 0)
    ws.sheet_view.showGridLines = False

    color_navy = "1F3864"
    color_blue = "2E75B6"
    color_light = "D6E4F0"
    color_white = "FFFFFF"
    color_gold = "C9A84C"
    color_green = "70AD47"
    color_orange = "ED7D31"
    color_gray = "F2F2F2"
    color_red = "C00000"

    def fill(color: str) -> PatternFill:
        return PatternFill("solid", fgColor=color)

    def font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
        return Font(bold=bold, color=color, size=size, name="맑은 고딕")

    center = Alignment(horizontal="center", vertical="center")
    left_center = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    for col, width in {
        1: 2, 2: 6, 3: 14, 4: 18, 5: 11, 6: 9, 7: 10, 8: 10, 9: 12, 10: 11, 11: 12, 12: 12,
        13: 3, 14: 18, 15: 10, 16: 10,
    }.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 23
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 8
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 10
    ws.row_dimensions[9].height = 8

    def style_range(cell_range: str, fill_color: str, font_obj: Font, alignment= None) -> None:
        alignment = alignment or center
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = fill(fill_color)
                cell.font = font_obj
                cell.alignment = alignment
                cell.border = border

    ws.merge_cells("B2:L2")
    style_range("B2:L2", color_navy, font(True, color_white, 18))
    ws["B2"] = "전종목 수급 스코어링 보고서"

    ws.merge_cells("B3:L3")
    style_range("B3:L3", color_blue, font(False, color_white, 10))
    ws["B3"] = f"기준일: {ref_date}  |  분석 대상: {len(final):,}개 종목  |  가격/수급/거래대금: {PRICE_WEIGHT:.0%}/{SUPPLY_WEIGHT:.0%}/{VALUE_WEIGHT:.0%}"

    grade_counts = {g: 0 for g in ["A", "B", "C", "D", "F"]}
    for grade, count in final["등급"].value_counts().items():
        if grade in grade_counts:
            grade_counts[grade] = int(count)

    score_65_plus = int((pd.to_numeric(final["종합스코어"], errors="coerce") >= 65).sum())
    score_80_plus = int((pd.to_numeric(final["종합스코어"], errors="coerce") >= 80).sum())
    cards = [
        ("전체 종목", len(final), color_blue),
        ("A등급", grade_counts["A"], color_gold),
        ("65점 이상", score_65_plus, color_green),
        ("80점 이상", score_80_plus, color_navy),
        ("수급 양호", int(final["기관외인동시순매수"].fillna(False).sum()), color_orange),
    ]
    for idx, (label, value, color) in enumerate(cards):
        col = 2 + idx * 2
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
        for cell in ws.iter_rows(min_row=6, max_row=6, min_col=col, max_col=col + 1):
            for c in cell:
                c.fill = fill(color)
                c.font = font(True, color_white, 9)
                c.alignment = center
                c.border = border
        for cell in ws.iter_rows(min_row=7, max_row=7, min_col=col, max_col=col + 1):
            for c in cell:
                c.fill = fill(color_gray)
                c.font = font(True, color, 18)
                c.alignment = center
                c.border = border

        lc = ws.cell(row=6, column=col)
        lc.value = label

        vc = ws.cell(row=7, column=col)
        vc.value = value

    ws.merge_cells("B10:L10")
    style_range("B10:L10", color_blue, font(True, color_white, 11), left_center)
    ws["B10"] = "[Top] 추천종목  (종합스코어 65점 이상)"
    ws["B10"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    headers = ["순위", "코드", "종목명", "종합", "등급", "가격", "수급", "거래대금", "5일수익률", "20일수익률", "기관/외인"]
    for offset, header in enumerate(headers, start=2):
        cell = ws.cell(row=11, column=offset)
        cell.value = header
        cell.font = font(True, color_white, 9)
        cell.fill = fill(color_navy)
        cell.alignment = center
        cell.border = border

    top = final[pd.to_numeric(final["종합스코어"], errors="coerce") >= 65].head(40).copy()
    grade_colors = {"A": color_gold, "B": color_green, "C": "FFC000", "D": color_orange, "F": "FF6B6B"}
    for idx, (_, row) in enumerate(top.iterrows(), start=1):
        excel_row = 11 + idx
        row_fill = fill(color_white if idx % 2 else color_gray)
        data = [
            idx,
            row.get("코드", ""),
            row.get("종목명", ""),
            row.get("종합스코어", ""),
            row.get("등급", ""),
            row.get("가격스코어", ""),
            row.get("수급스코어", ""),
            row.get("거래대금스코어", ""),
            row.get("5일수익률(%)", ""),
            row.get("20일수익률(%)", ""),
            "동시순매수" if bool(row.get("기관외인동시순매수", False)) else "",
        ]
        for offset, value in enumerate(data, start=2):
            cell = ws.cell(row=excel_row, column=offset)
            cell.value = value
            cell.font = font(False, "000000", 9)
            cell.fill = row_fill
            cell.alignment = center
            cell.border = border
        grade_cell = ws.cell(row=excel_row, column=6)
        grade = str(row.get("등급", "-"))
        grade_cell.fill = fill(grade_colors.get(grade, "BFBFBF"))
        grade_cell.font = font(True, color_white, 9)

    chart_row = 10
    chart_col_label = 14
    chart_col_value = 15
    ws.cell(row=chart_row, column=chart_col_label).value = "구분"
    ws.cell(row=chart_row, column=chart_col_value).value = "종목수"
    chart_data = [
        ("강세 (A+B)", grade_counts["A"] + grade_counts["B"]),
        ("중립 (C)", grade_counts["C"]),
        ("약세 (D+F)", grade_counts["D"] + grade_counts["F"]),
    ]
    for idx, (label, value) in enumerate(chart_data, start=1):
        ws.cell(row=chart_row + idx, column=chart_col_label).value = label
        ws.cell(row=chart_row + idx, column=chart_col_value).value = value

    for row in range(chart_row, chart_row + 4):
        for col in [chart_col_label, chart_col_value]:
            cell = ws.cell(row=row, column=col)
            cell.font = font(True if row == chart_row else False, color_white if row == chart_row else "000000", 9)
            cell.fill = fill(color_navy if row == chart_row else color_light)
            cell.alignment = center
            cell.border = border
    ws.column_dimensions["N"].hidden = True
    ws.column_dimensions["O"].hidden = True

    pie = PieChart()
    pie.title = "전종목 스코어 분포"
    pie.style = 10
    data_ref = Reference(ws, min_col=chart_col_value, min_row=chart_row, max_row=chart_row + 3)
    label_ref = Reference(ws, min_col=chart_col_label, min_row=chart_row + 1, max_row=chart_row + 3)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(label_ref)
    for idx, color in enumerate([color_blue, color_orange, color_red]):
        point = DataPoint(idx=idx)
        point.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(point)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.width = 12
    pie.height = 9
    ws.add_chart(pie, "N15")


def save_outputs(scored: pd.DataFrame, history: pd.DataFrame, sheet: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = OUTPUT_DIR / f"전종목_수급_스코어링_{datetime.now().strftime('%Y%m%d')}.xlsx"

    output_cols = [
        "date", "code", "name", "total_score", "grade",
        "price_score", "supply_score", "value_score",
        "close", "ret_5d", "ret_20d", "ma5", "ma10", "ma20",
        "above_ma10", "ma5_above_ma10",
        "turnover_avg_5d", "turnover_avg_20d", "turnover_rvol",
        "inst_net_5d", "inst_net_20d", "foreign_net_5d", "foreign_net_20d",
        "inst_ratio_5d", "inst_ratio_20d", "foreign_ratio_5d", "foreign_ratio_20d",
        "inst_positive_rate_20d", "foreign_positive_rate_20d", "both_supply_positive",
    ]
    for col in output_cols:
        if col not in scored.columns:
            scored[col] = pd.NA

    rename = {
        "date": "기준일",
        "code": "코드",
        "name": "종목명",
        "total_score": "종합스코어",
        "grade": "등급",
        "price_score": "가격스코어",
        "supply_score": "수급스코어",
        "value_score": "거래대금스코어",
        "close": "종가",
        "ret_5d": "5일수익률(%)",
        "ret_20d": "20일수익률(%)",
        "ma5": "5일선",
        "ma10": "10일선",
        "ma20": "20일선",
        "above_ma10": "10일선상회",
        "ma5_above_ma10": "5일선>10일선",
        "turnover_avg_5d": "거래대금5일평균(원)",
        "turnover_avg_20d": "거래대금20일평균(원)",
        "turnover_rvol": "거래대금5일/20일",
        "inst_net_5d": "기관5일순매수(만원)",
        "inst_net_20d": "기관20일순매수(만원)",
        "foreign_net_5d": "외국인5일순매수(만원)",
        "foreign_net_20d": "외국인20일순매수(만원)",
        "inst_ratio_5d": "기관5일순매수/거래대금(%)",
        "inst_ratio_20d": "기관20일순매수/거래대금(%)",
        "foreign_ratio_5d": "외국인5일순매수/거래대금(%)",
        "foreign_ratio_20d": "외국인20일순매수/거래대금(%)",
        "inst_positive_rate_20d": "기관수급양수비율20일(%)",
        "foreign_positive_rate_20d": "외국인수급양수비율20일(%)",
        "both_supply_positive": "기관외인동시순매수",
    }

    final = scored[output_cols].rename(columns=rename)
    numeric_cols = final.select_dtypes(include=["number"]).columns
    final[numeric_cols] = final[numeric_cols].round(2)
    recommend = final[pd.to_numeric(final["종합스코어"], errors="coerce") >= 65].copy()

    summary = pd.DataFrame(
        [
            {"항목": "입력파일", "값": str(INPUT_FILE)},
            {"항목": "사용시트", "값": sheet},
            {"항목": "기준일", "값": scored["date"].max().strftime("%Y-%m-%d")},
            {"항목": "대상종목수", "값": len(final)},
            {"항목": "65점 이상", "값": len(recommend)},
            {"항목": "가격/수급/거래대금 가중치", "값": f"{PRICE_WEIGHT:.0%}/{SUPPLY_WEIGHT:.0%}/{VALUE_WEIGHT:.0%}"},
        ]
    )

    a_grade = final[final["등급"] == "A"].copy()
    foreign_top = final.sort_values("외국인5일순매수/거래대금(%)", ascending=False, na_position="last").head(50)
    inst_top = final.sort_values("기관5일순매수/거래대금(%)", ascending=False, na_position="last").head(50)
    value_top = final.sort_values("거래대금5일/20일", ascending=False, na_position="last").head(50)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _create_report_sheet(writer, final, scored, scored["date"].max().strftime("%Y-%m-%d"))
        recommend.to_excel(writer, sheet_name="추천종목", index=False)
        a_grade.to_excel(writer, sheet_name="A등급", index=False)
        final.to_excel(writer, sheet_name="전체", index=False)
        foreign_top.to_excel(writer, sheet_name="외국인순매수TOP", index=False)
        inst_top.to_excel(writer, sheet_name="기관순매수TOP", index=False)
        value_top.to_excel(writer, sheet_name="거래대금TOP", index=False)
        summary.to_excel(writer, sheet_name="요약", index=False)

        for sheet_name in ["추천종목", "A등급", "전체", "외국인순매수TOP", "기관순매수TOP", "거래대금TOP", "요약"]:
            _format_table_sheet(writer.sheets[sheet_name])

    return output


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(INPUT_FILE)

    wide, sheet = load_workbook_data(INPUT_FILE)
    history = add_time_series_features(wide)
    scored = latest_scoring_frame(history)
    output = save_outputs(scored, history, sheet)

    top_cols = ["date", "code", "name", "total_score", "grade", "price_score", "supply_score", "value_score"]
    print("\n[TOP 20]")
    print(scored[top_cols].head(20).to_string(index=False))
    print(f"\n저장 완료: {output}")


if __name__ == "__main__":
    main()
